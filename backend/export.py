"""
Excel export functionality using openpyxl.
"""

import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models import JobListing
from config import EXPORT_DIR


def generate_excel(jobs: list[JobListing]) -> BytesIO:
    """
    Generate a styled Excel file from a list of job listings.
    Returns a BytesIO buffer ready for streaming.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Indeed PK Jobs"

    # ── Column headers ─────────────────────────────────────────────
    headers = [
        "Job Title",
        "Company",
        "Location",
        "Date Posted",
        "Job Type",
        "Description",
        "Apply Link",
        "Source URL",
        "Scraped At",
    ]

    # ── Styles ─────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        bottom=Side(style="medium", color="2C5F8A"),
    )

    cell_font = Font(name="Calibri", size=11)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    alt_row_fill = PatternFill(start_color="F0F5FB", end_color="F0F5FB", fill_type="solid")

    # ── Write headers ──────────────────────────────────────────────
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # ── Write data rows ────────────────────────────────────────────
    for row_idx, job in enumerate(jobs, 2):
        row_data = [
            job.title,
            job.company,
            job.location,
            job.date_posted,
            job.job_type,
            job.description[:5000] if job.description else "",  # Cap at 5000 chars for Excel
            job.apply_link,
            job.source_url,
            job.scraped_at or "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_alignment

            # Alternate row coloring
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill

    # ── Column widths ──────────────────────────────────────────────
    column_widths = [30, 25, 20, 15, 15, 60, 40, 50, 22]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze header row ──────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Auto-filter ────────────────────────────────────────────────
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(jobs) + 1}"

    # ── Save to buffer ─────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def save_excel_to_disk(jobs: list[JobListing], filename: str = "indeed_jobs.xlsx") -> str:
    """Save Excel file to disk and return the file path."""
    os.makedirs(EXPORT_DIR, exist_ok=True)
    filepath = os.path.join(EXPORT_DIR, filename)

    buffer = generate_excel(jobs)
    with open(filepath, "wb") as f:
        f.write(buffer.getvalue())

    return filepath
